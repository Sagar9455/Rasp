from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "/home/mobase/Rasp/Sahithi/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/home/mobase/Rasp/Sahithi/template_DiagserviceList.xlsx"
output_excel_file = "/home/mobase/Rasp/Sahithi/output/ouput_DiagserviceList_01.xlsx"

# === Step 1: Parse the CDD file ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# Load all lines as strings for reference lookups
with open(input_cdd_file, 'r', encoding='utf-8') as f:
    all_text_lines = f.readlines()

# === Step 2: Prepare data extraction ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # === Step 3: Extract Service ID ===
        # Step 3.1: Find tmplref from <SERVICE>
        service_elem = diaginst.find(".//SERVICE")
        tmplref_v1 = service_elem.get("tmplref") if service_elem is not None else ""

        # Step 3.2: Find V2 from <DCLSRVTMPL id=V1 tmplref=V2>
        tmplref_v2 = ""
        for line in all_text_lines:
            if f'<DCLSRVTMPL id="{tmplref_v1}"' in line:
                pos = line.find("tmplref=")
                if pos != -1:
                    tmplref_v2 = line[pos:].split('"')[1]
                break

        # Step 3.3: Find Service ID from <PROTOCOLSERVICE id=V2>
        service_id = ""
        for idx, line in enumerate(all_text_lines):
            if f'<PROTOCOLSERVICE id="{tmplref_v2}"' in line:
                # Look at the next line for the TUV content
                if idx + 1 < len(all_text_lines):
                    next_line = all_text_lines[idx + 1]
                    if "TUV" in next_line and "($" in next_line:
                        try:
                            # Extract between '>($' and ')'
                            start = next_line.index(">$(") + 3
                            end = next_line.index(")", start)
                            service_id = next_line[start:end]
                        except ValueError:
                            pass
                break

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 4: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active

start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"     # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                    # Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                      # Service ID
    ws.cell(row=i, column=4).value = row_data[0]                      # Service Name
    ws.cell(row=i, column=5).value = row_data[2]                      # SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                      # SubService Name

# === Save final Excel ===
wb.save(output_excel_file)
print(f"✅ Excel saved as: {output_excel_file}")
