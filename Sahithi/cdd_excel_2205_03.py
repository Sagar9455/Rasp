from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "your_input_file.cdd"  # Replace with actual CDD path
template_file = "your_template_file.xlsx"  # Replace with actual template path
output_excel_file = "template_DiagserviceList_filled.xlsx"

# === Step 1: Parse the XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# === Step 2: Build lookup maps from XML ===

# Map: DCLSRVTMPL id -> tmplref (e.g. "_0C864D20" -> "_04AB2210")
dclsrvtmpl_map = {}
for dcl in root.xpath(".//DCLSRVTMPL"):
    id_val = dcl.get("id", "")
    tmplref_val = dcl.get("tmplref", "")
    if id_val and tmplref_val:
        dclsrvtmpl_map[id_val] = tmplref_val

# Map: PROTOCOLSERVICE id -> service ID (value inside <TUV>)
protocol_map = {}
for proto in root.xpath(".//PROTOCOLSERVICE"):
    proto_id = proto.get("id", "")
    tuv_elem = proto.find(".//TUV")
    if proto_id and tuv_elem is not None and tuv_elem.text:
        text = tuv_elem.text.strip()
        # Expecting something like: >($10)
        if "$" in text:
            try:
                start = text.index("$") + 1
                end = text.index(")", start)
                service_id = text[start:end]
                protocol_map[proto_id] = service_id
            except ValueError:
                continue

# === Step 3: Extract DIAGCLASS content ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.find(".//STATICVALUE")
        if staticvalue_elem is not None:
            value = staticvalue_elem.get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # === Step 4: Get service ID through reference chain ===
        tmplref_v1 = ""
        service_elem = diaginst.find(".//SERVICE")
        if service_elem is not None:
            tmplref_v1 = service_elem.get("tmplref", "")

        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        service_id = protocol_map.get(tmplref_v2, "")

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 5: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                 # Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                   # Service ID
    ws.cell(row=i, column=4).value = row_data[0]                   # Service Name
    ws.cell(row=i, column=5).value = row_data[2]                   # SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                   # SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
