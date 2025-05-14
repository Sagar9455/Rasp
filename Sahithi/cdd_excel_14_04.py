from lxml import etree
from openpyxl import load_workbook
import os

# === Input Files ===
input_cdd_file = "/home/mobase/Rasp/Sahithi/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/home/mobase/Rasp/Sahithi/template_DiagserviceList.xlsx"
output_excel_file = "/home/mobase/Rasp/Sahithi/output/ouput_DiagserviceList_05.xlsx"

# Extract file name for Excel column
cdd_file_name = os.path.basename(input_cdd_file)

# === Parse XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()

# Build lookup maps
service_id_to_tmplref = {}
tmplref1_to_tmplref2 = {}
tmplref2_to_service_id_str = {}

# 1. Map <SERVICE id=... tmplref=V1>
for service in root.xpath(".//SERVICE"):
    sid = service.get("id")
    tmplref = service.get("tmplref")
    if sid and tmplref:
        service_id_to_tmplref[sid] = tmplref

# 2. Map <DCLSRVTMPL id=V1 tmplref=V2>
for tmpl1 in root.xpath(".//DCLSRVTMPL"):
    tid = tmpl1.get("id")
    tmplref = tmpl1.get("tmplref")
    if tid and tmplref:
        tmplref1_to_tmplref2[tid] = tmplref

# 3. Map <PROTOCOLSERVICE id=V2> to serviceId string from <TUV>
for psvc in root.xpath(".//PROTOCOLSERVICE"):
    pid = psvc.get("id")
    tuv = psvc.xpath(".//TUV")
    if pid and tuv:
        text = tuv[0].text
        if text and "($" in text and ")" in text:
            service_id_str = text.split("($")[-1].split(")")[0]
            tmplref2_to_service_id_str[pid] = service_id_str

# === Collect DIAGINST Data ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # Step 1: SERVICE id
        parent_service = diaginst.getparent()
        while parent_service is not None and parent_service.tag != "SERVICE":
            parent_service = parent_service.getparent()

        service_id_str = ""
        if parent_service is not None:
            service_id = parent_service.get("id")
            if service_id in service_id_to_tmplref:
                tmplref1 = service_id_to_tmplref[service_id]
                tmplref2 = tmplref1_to_tmplref2.get(tmplref1, "")
                service_id_str = tmplref2_to_service_id_str.get(tmplref2, "")

        results.append([service_name, subservice_name, static_value_hex, service_id_str])

# === Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"        # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                       # Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                         # Service ID
    ws.cell(row=i, column=4).value = row_data[0]                         # Service Description
    ws.cell(row=i, column=5).value = row_data[2]                         # SubService ID (Hex)
    ws.cell(row=i, column=6).value = row_data[1]                         # SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved as: {output_excel_file}")
