from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "/home/mobase/Rasp/Sahithi/backup/input/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/home/mobase/Rasp/Sahithi/backup/input/template_DiagserviceList.xlsx"
output_excel_file = "/home/mobase/Rasp/Sahithi/backup/output/MKBD_DiagserviceList_2305.xlsx"

# === Step 1: Parse the XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# === Step 2: Build lookup maps from XML ===

# Map: DCLSRVTMPL id -> tmplref (e.g. "_0C864D20" -> "_04AB2210")
dclsrvtmpl_map = {}
for dcl in root.xpath(".//DCLSRVTMPL"):
    id_val = dcl.get("id", "")
    tmplref_val = dcl.get("tmplref", "")
    if id_val and tmplref_val:
        dclsrvtmpl_map[id_val] = tmplref_val

# Map: PROTOCOLSERVICE id -> service ID (value inside <TUV>)
protocol_map = {}
for proto in root.xpath(".//PROTOCOLSERVICE"):
    proto_id = proto.get("id", "")
    tuv_elem = proto.find(".//TUV")
    if proto_id and tuv_elem is not None and tuv_elem.text:
        text = tuv_elem.text.strip()
        # Expecting something like: >($10)
        if "$" in text:
            try:
                start = text.index("$") + 1
                end = text.index(")", start)
                service_id = text[start:end]
                protocol_map[proto_id] = service_id
            except ValueError:
                continue

# === Step 3: Extract DIAGCLASS content ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.find(".//STATICVALUE")
        if staticvalue_elem is not None:
            value = staticvalue_elem.get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # === Step 4: Get service ID through reference chain ===
        tmplref_v1 = ""
        service_elem = diaginst.find(".//SERVICE")
        if service_elem is not None:
            tmplref_v1 = service_elem.get("tmplref", "")

        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        raw_id = protocol_map.get(tmplref_v2, "")
        service_id = f"0x{raw_id}"if raw_id else ""
        
         
        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 5: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    service_id = row_data[3]
    pos_response = ""

    # Compute Positive Response: service ID + 0x40
    if service_id.startswith("0x"):
        try:
            pos_response = f"0x{int(service_id, 16) + 0x40:02X}"
        except ValueError:
            pass
    
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # Column A: TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                 # Column B: Ref Document
    ws.cell(row=i, column=3).value = service_id                    # Column C: Service ID
    ws.cell(row=i, column=4).value = row_data[0]                   # Column D: Service Name
    ws.cell(row=i, column=5).value = row_data[2]                   # Column E: SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                   # Column F: SubService Name
    ws.cell(row=i, column=7).value = pos_response                  # Column G: Positive Response

wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
