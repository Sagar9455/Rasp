from lxml import etree
from openpyxl import load_workbook
import os
import re

# === Input Files ===
input_cdd_file = "/home/mobase/Rasp/Sahithi/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/home/mobase/Rasp/Sahithi/template_DiagserviceList.xlsx"
output_excel_file = "/home/mobase/Rasp/Sahithi/output/ouput_DiagserviceList_03.xlsx"

# === Setup ===
cdd_file_name = os.path.basename(input_cdd_file)
tree = etree.parse(input_cdd_file)
root = tree.getroot()

# === Step 1: Pre-parse <SERVICE>, <DCLSRVTMPL>, and <PROTOCOLSERVICE> entries ===

# SERVICE id -> tmplref
service_tmplref_map = {
    service.get("id"): service.get("tmplref")
    for service in root.xpath(".//SERVICE") if service.get("tmplref")
}

# DCLSRVTMPL id -> tmplref
dclsrvtmpl_map = {
    dcl.get("id"): dcl.get("tmplref")
    for dcl in root.xpath(".//DCLSRVTMPL") if dcl.get("tmplref")
}

# PROTOCOLSERVICE id -> service ID (text from <TUV> after "($" and before ")")
protocol_service_id_map = {}
for ps in root.xpath(".//PROTOCOLSERVICE"):
    ps_id = ps.get("id")
    parent = ps.getparent()
    idx = parent.index(ps)
    # Look for <TUV> after <PROTOCOLSERVICE>
    for offset in range(1, 4):  # check next few siblings
        if idx + offset < len(parent):
            elem = parent[idx + offset]
            if elem.tag == "TUV" and elem.text:
                match = re.search(r"\(\$(.*?)\)", elem.text)
                if match:
                    protocol_service_id_map[ps_id] = match.group(1)
                    break

# === Step 2: Extract all diagnostic entries ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID (hex)
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # Find SERVICE id ref from DIAGINST
        serviceref = diaginst.get("serviceref")
        service_id = ""
        if serviceref:
            v1 = service_tmplref_map.get(serviceref)
            if v1:
                v2 = dclsrvtmpl_map.get(v1)
                if v2:
                    service_id = protocol_service_id_map.get(v2, "")

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 3: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"      # Column 1: TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                     # Column 2: Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                       # Column 3: Service ID (from PROTOCOLSERVICE)
    ws.cell(row=i, column=4).value = row_data[0]                       # Column 4: Service Name
    ws.cell(row=i, column=5).value = row_data[2]                       # Column 5: SubService ID (hex)
    ws.cell(row=i, column=6).value = row_data[1]                       # Column 6: SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved as: {output_excel_file}")
