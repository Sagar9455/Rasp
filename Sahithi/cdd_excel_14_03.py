from lxml import etree
import os
import re
from openpyxl import load_workbook

# === Input Files ===
input_cdd_file = "/path/to/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/path/to/template_with_checkboxes.xlsx"
output_excel_file = "template_DiagserviceList.xlsx"

# Extract filename
cdd_file_name = os.path.basename(input_cdd_file)

# === Parse the XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()

# === Build lookup tables ===
# Step 1: SERVICE id → tmplref
service_id_to_tmplref = {
    elem.get("id"): elem.get("tmplref")
    for elem in root.xpath(".//SERVICE") if elem.get("id") and elem.get("tmplref")
}

# Step 2: DCLSRVTMPL id → tmplref
dclsrvtmpl_id_to_tmplref = {
    elem.get("id"): elem.get("tmplref")
    for elem in root.xpath(".//DCLSRVTMPL") if elem.get("id") and elem.get("tmplref")
}

# Step 3: PROTOCOLSERVICE id → service ID from TUV (e.g., extract text inside ($xx))
protocolservice_id_to_service_id = {}
for ps in root.xpath(".//PROTOCOLSERVICE"):
    ps_id = ps.get("id")
    parent = ps.getparent()
    index = list(parent).index(ps)
    # Try next 1-3 siblings to find <TUV>
    for i in range(1, 4):
        if index + i < len(parent):
            next_elem = parent[index + i]
            if next_elem.tag == "TUV" and next_elem.text:
                match = re.search(r"\(\$(.*?)\)", next_elem.text)
                if match:
                    protocolservice_id_to_service_id[ps_id] = match.group(1)
                    break

# === Collect final service data ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # Get SubService ID (STATICVALUE)
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # Step 1: Get SERVICE id reference from diaginst
        service_ref_elem = diaginst.find(".//SERVICE")
        service_id = service_ref_elem.get("id") if service_ref_elem is not None else None

        service_id_final = ""
        if service_id and service_id in service_id_to_tmplref:
            v1 = service_id_to_tmplref[service_id]  # tmplref from <SERVICE>
            v2 = dclsrvtmpl_id_to_tmplref.get(v1)   # tmplref from <DCLSRVTMPL>
            service_id_final = protocolservice_id_to_service_id.get(v2, "")

        results.append([service_name, subservice_name, service_id_final, static_value_hex])

# === Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"   # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                  # Ref Document
    ws.cell(row=i, column=3).value = row[2]                         # Service ID (new logic)
    ws.cell(row=i, column=4).value = row[0]                         # Service Name
    ws.cell(row=i, column=5).value = row[3]                         # SubService ID Hex
    ws.cell(row=i, column=6).value = row[1]                         # SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
