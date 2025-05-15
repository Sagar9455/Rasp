from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "your_input_file.cdd"
template_file = "your_template_file.xlsx"
output_excel_file = "template_DiagserviceList_filled.xlsx"

# === Step 1: Parse the XML ===
parser = etree.XMLParser(load_dtd=False, no_network=True, recover=True)
tree = etree.parse(input_cdd_file, parser)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# === Step 2: Build lookup maps from XML ===

# Map: DCLSRVTMPL id → tmplref (points to PROTOCOLSERVICE id)
dclsrvtmpl_map = {
    e.get("id"): e.get("tmplref")
    for e in root.xpath(".//DCLSRVTMPL[@id][@tmplref]")
}

# Map: PROTOCOLSERVICE id → service ID from TUV (e.g., ($10))
protocol_map = {}
for e in root.xpath(".//PROTOCOLSERVICE[@id]"):
    proto_id = e.get("id")
    tuv_elem = e.find(".//TUV")
    if tuv_elem is not None and tuv_elem.text and "($" in tuv_elem.text:
        try:
            service_id = tuv_elem.text.split("($")[1].split(")")[0]
            protocol_map[proto_id] = service_id
        except IndexError:
            continue

# === Step 3: Extract DIAGCLASS and DIAGINST content ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID (STATICVALUE.v as hex)
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # Step 4: Get service ID through reference chain
        service_elem = diaginst.find(".//SERVICE")
        tmplref_v1 = service_elem.get("tmplref") if service_elem is not None else ""
        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        service_id = protocol_map.get(tmplref_v2, "")

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 5: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                 # Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                   # Service ID
    ws.cell(row=i, column=4).value = row_data[0]                   # Service Name
    ws.cell(row=i, column=5).value = row_data[2]                   # SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                   # SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
