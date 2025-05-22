import os
import re
from lxml import etree
from openpyxl import load_workbook

# === Input files ===
input_cdd_file = "your_input_file.cdd"
template_file = "your_template_file.xlsx"
output_excel_file = "template_DiagserviceList_filled.xlsx"

# === Step 1: Parse the CDD XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# Load CDD file as text lines
with open(input_cdd_file, 'r', encoding='utf-8') as f:
    all_text_lines = f.readlines()

# === Step 2: Build reference maps ===

# Map: DCLSRVTMPL id -> tmplref
dclsrvtmpl_map = {}
for line in all_text_lines:
    match = re.search(r'<DCLSRVTMPL\s+id=["\']([^"\']+)["\']\s+tmplref=["\']([^"\']+)["\']', line)
    if match:
        dclsrvtmpl_map[match.group(1)] = match.group(2)

# Map: PROTOCOLSERVICE id -> Service ID from TUV
protocol_map = {}
for i, line in enumerate(all_text_lines):
    match = re.search(r'<PROTOCOLSERVICE\s+id=["\']([^"\']+)["\']', line)
    if match:
        proto_id = match.group(1)
        next_line = all_text_lines[i + 1] if i + 1 < len(all_text_lines) else ""
        tuv_match = re.search(r'>\(\$([^)]+)\)', next_line)
        if tuv_match:
            protocol_map[proto_id] = tuv_match.group(1)

# === Step 3: Extract service data ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # === Reference Chain to Get Service ID ===
        service_elem = diaginst.xpath(".//SERVICE")
        tmplref_v1 = service_elem[0].get("tmplref") if service_elem else ""
        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        service_id = protocol_map.get(tmplref_v2, "")

        # Debug print (optional)
        # print(f"[DEBUG] {service_name} → {subservice_name} | V1: {tmplref_v1} → V2: {tmplref_v2} → ServiceID: {service_id}")

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 4: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                 # Ref Document
    ws.cell(row=i, column=3).value = row_data[3]                   # Service ID
    ws.cell(row=i, column=4).value = row_data[0]                   # Service Name
    ws.cell(row=i, column=5).value = row_data[2]                   # SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                   # SubService Name

wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
