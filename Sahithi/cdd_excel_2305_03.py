from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "/mnt/data/08225c9e-ac96-4067-a67a-b32e0d8149fa.cdd"
template_file = "/mnt/data/ebf09fbf-1591-41fa-ac15-7468cfbcb30d.xlsx"
output_excel_file = "/mnt/data/output_diag_service_list.xlsx"

# === Step 1: Parse the XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# === Step 2: Build lookup maps ===

# Map: DCLSRVTMPL id -> tmplref
dclsrvtmpl_map = {}
for dcl in root.xpath(".//DCLSRVTMPL"):
    id_val = dcl.get("id", "")
    tmplref_val = dcl.get("tmplref", "")
    if id_val and tmplref_val:
        dclsrvtmpl_map[id_val] = tmplref_val

# Map: PROTOCOLSERVICE id -> service ID (from TUV)
protocol_map = {}
for proto in root.xpath(".//PROTOCOLSERVICE"):
    proto_id = proto.get("id", "")
    tuv_elem = proto.find(".//TUV")
    if proto_id and tuv_elem is not None and tuv_elem.text:
        text = tuv_elem.text.strip()
        if "$" in text:
            try:
                start = text.index("$") + 1
                end = text.index(")", start)
                service_id = text[start:end]
                protocol_map[proto_id] = service_id
            except ValueError:
                continue

# === Step 3: Extract data from DIAGCLASS/DIAGINST ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        # SubService Name from <SHORTCUTQUAL>
        shortcut_elem = diaginst.find(".//SHORTCUTQUAL")
        subservice_name = shortcut_elem.text.strip() if shortcut_elem is not None else ""

        # SubService ID from STATICVALUE
        static_value_hex = ""
        staticvalue_elem = diaginst.find(".//STATICVALUE")
        if staticvalue_elem is not None:
            value = staticvalue_elem.get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # Find Service ID through tmplref chains
        service_elem = diaginst.find(".//SERVICE")
        tmplref_v1 = service_elem.get("tmplref", "") if service_elem is not None else ""
        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        raw_service_id = protocol_map.get(tmplref_v2, "")
        service_id = f"0x{raw_service_id}" if raw_service_id else ""

        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 4: Write to Excel Template ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    service_id = row_data[3]
    pos_response = ""

    # Compute Positive Response = Service ID + 0x40
    if service_id.startswith("0x"):
        try:
            pos_response = f"0x{int(service_id, 16) + 0x40:02X}"
        except ValueError:
            pass

    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # A
    ws.cell(row=i, column=2).value = cdd_file_name                 # B
    ws.cell(row=i, column=3).value = service_id                    # C
    ws.cell(row=i, column=4).value = row_data[0]                   # D
    ws.cell(row=i, column=5).value = row_data[2]                   # E
    ws.cell(row=i, column=6).value = row_data[1]                   # F
    ws.cell(row=i, column=7).value = pos_response                  # G

wb.save(output_excel_file)
print(f"✅ Excel saved to: {output_excel_file}")
