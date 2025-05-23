from lxml import etree
from openpyxl import load_workbook
import os

# === Input files ===
input_cdd_file = "/home/mobase/Rasp/Sahithi/KY_MKBD_Diagnostic_Rev01.cdd"
template_file = "/home/mobase/Rasp/Sahithi/template_DiagserviceList.xlsx"
output_excel_file = "MKBD_DiagserviceList_2305_01.xlsx"

# === Step 1: Parse the XML ===
tree = etree.parse(input_cdd_file)
root = tree.getroot()
cdd_file_name = os.path.basename(input_cdd_file)

# Read raw lines for text-based matching
with open(input_cdd_file, 'r', encoding='utf-8') as f:
    all_text_lines = f.readlines()

# === Step 2: Build lookup maps ===

# Map: DCLSRVTMPL id -> tmplref
dclsrvtmpl_map = {}
for line in all_text_lines:
    if "<DCLSRVTMPL " in line and "id=" in line and "tmplref=" in line:
        try:
            id_val = line.split('id="')[1].split('"')[0]
            tmplref_val = line.split('tmplref="')[1].split('"')[0]
            dclsrvtmpl_map[id_val] = tmplref_val
        except IndexError:
            continue

# Map: PROTOCOLSERVICE id -> service ID (extracted from TUV >($xx)
protocol_map = {}
for i, line in enumerate(all_text_lines):
    if "<PROTOCOLSERVICE " in line and "id=" in line:
        try:
            proto_id = line.split('id="')[1].split('"')[0]
            next_line = all_text_lines[i + 1]
            if "TUV" in next_line and "($" in next_line:
                start = next_line.index(">$(") + 3
                end = next_line.index(")", start)
                service_id = next_line[start:end]
                if service_id:
                    service_id = f"0x{service_id}"  # Add 0x prefix here
                    protocol_map[proto_id] = service_id
        except (IndexError, ValueError):
            continue

# === Step 3: Extract DIAGCLASS content ===
results = []

for diagclass in root.xpath(".//DIAGCLASS"):
    service_name_elem = diagclass.find(".//QUAL")
    service_name = service_name_elem.text.strip() if service_name_elem is not None else ""

    for diaginst in diagclass.xpath(".//DIAGINST"):
        subservice_name_elem = diaginst.find(".//QUAL")
        subservice_name = subservice_name_elem.text.strip() if subservice_name_elem is not None else ""

        # SubService ID
        static_value_hex = ""
        staticvalue_elem = diaginst.xpath(".//STATICVALUE")
        if staticvalue_elem:
            value = staticvalue_elem[0].get("v", "")
            if value.isdigit():
                static_value_hex = f"0x{int(value):02X}"

        # === Step 4: Get service ID through reference chain ===
        service_elem = diaginst.find(".//SERVICE")
        tmplref_v1 = service_elem.get("tmplref") if service_elem is not None else ""
        tmplref_v2 = dclsrvtmpl_map.get(tmplref_v1, "")
        service_id = protocol_map.get(tmplref_v2, "")
        print(f"✅ Excel saved service_id: {service_id}")
        results.append([service_name, subservice_name, static_value_hex, service_id])

# === Step 5: Write to Excel ===
wb = load_workbook(template_file)
ws = wb.active
start_row = 2

for i, row_data in enumerate(results, start=start_row):
    service_id = row_data[3]
    pos_response = ""

    # Compute Positive Response: service ID + 0x40
    if service_id.startswith("0x"):
        try:
            pos_response = f"0x{int(service_id, 16) + 0x40:02X}"
        except ValueError:
            pass
    print(f"✅ Excel saved pos_response: {pos_response}")
    
    ws.cell(row=i, column=1).value = f"TC_ID-{i - start_row + 1}"  # Column A: TC_ID
    ws.cell(row=i, column=2).value = cdd_file_name                 # Column B: Ref Document
    ws.cell(row=i, column=3).value = service_id                    # Column C: Service ID
    ws.cell(row=i, column=4).value = row_data[0]                   # Column D: Service Name
    ws.cell(row=i, column=5).value = row_data[2]                   # Column E: SubService ID
    ws.cell(row=i, column=6).value = row_data[1]                   # Column F: SubService Name
    ws.cell(row=i, column=7).value = pos_response                  # Column G: Positive Response

# === Save Excel ===
wb.save(output_excel_file)
print(f"✅ Excel saved: {output_excel_file}")
